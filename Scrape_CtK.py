from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import *
from selenium.webdriver.chrome.options import Options
from bs4 import BeautifulSoup
import requests
import time
from openpyxl import load_workbook
from openpyxl.styles import Alignment
import datetime
import os

# Start time of program
start_time = time.time()

# Create the payload
usernameStr = 'Meatball'
passwordStr = 'Seeker666*'

# Global variables
countryList = []
country = []
coinCC = []
coinG = []

# Chrome
# Change the option to start maximized
options = webdriver.ChromeOptions()
options.add_argument("--window-size=1600,1000")
options.add_argument("--headless")
# Mac
# browser = webdriver.Chrome('/Users/lucian/Google Drive/Python/chromedriver', options=options)
# Work
browser = webdriver.Chrome('chromedriver.exe', options=options)

# Get the site
browser.get("https://www.cavetokingdom.com/")

# Post the payload to the site to log in
username = browser.find_element_by_xpath('//input[@name="uname"]')
username.send_keys(usernameStr)
password = browser.find_element_by_xpath('//input[@name="pass"]')
password.send_keys(passwordStr)
signInButton = browser.find_element_by_xpath('//input[@name="login"]')
signInButton.click()

# Get the page of partners
browser.get('https://www.cavetokingdom.com/account/partners')
# List of all countries
countries = browser.find_elements_by_xpath('//div[@class="_ms_country_nr"]')
# For each country, it searches, clicks on it and then saves the data from it
for i in range(len(countries)):
    countries = browser.find_elements_by_xpath('//div[@class="_ms_country_nr"]')
    elem = countries[i].get_attribute('innerHTML').replace('\n', '')
    countryList.append(elem[:-1])
    WebDriverWait(browser, 30).until(EC.element_to_be_clickable((By.XPATH, '//div[@class="_ms_buy_share_text"]')))
    browser.find_element_by_xpath('//div[@class="_ms_buy_share_text"]').click()
    browser.find_element_by_id('_ms_search_country').send_keys(countryList[i])
    time.sleep(1)
    countries[i].click()
    WebDriverWait(browser, 30).until(EC.text_to_be_present_in_element(
        (By.XPATH, '//div[@class="_ms_buy_share_text"]'), countries[i].text))
    partners = browser.execute_script('return document.documentElement.outerHTML')
    soup = BeautifulSoup(partners, 'html.parser')
    countryName = soup.find("div", attrs={"class": "_ms_buy_share_text"})
    country.append(countryName.text)
    print(countryName.text)
    coinValues = soup.find_all("div", attrs={"class": "_ms_your_coins_nr"})
    tmp = coinValues[0].text.replace(' CC', '')
    coinCC.append(tmp.replace(',', ''))
    tmp1 = coinValues[1].text.replace(' Gold', '')
    coinG.append(tmp1.replace(',', ''))

today = datetime.datetime.now().strftime('%d.%m.%Y')
wb = load_workbook('CtK.xlsx')
ws = wb.active

i = 0
for row in ws.iter_rows(min_row=1, min_col=1, max_col=61, max_row=253):
    counter = 0
    for cell in reversed(row):
        counter += 1
        if 2 < counter < 61:
            if counter % 2 == 0:
                ws[str(cell2).replace('>', '')[17:]].value = cell.value
            else:
                if cell.value is not None:
                    ws[str(cell1).replace('>', '')[17:]].value = cell.value
            if counter == 59:
                if str(cell) != "<MergedCell 'Tabelle1'.C1>":
                    cell.value = float(coinG[i])
            elif counter == 60:
                if str(cell).replace('>', '')[17:] == 'B1':
                    cell.value = today
                else:
                    cell.value = float(coinCC[i])
        elif counter == 61:
            if cell.value is not None:
                cell.value = country[i]
                i += 1
        if counter % 2 == 0:
            cell2 = cell
        else:
            cell1 = cell

wb.save('CtK.xlsx')

print("--- %s seconds ---" % (time.time() - start_time))
